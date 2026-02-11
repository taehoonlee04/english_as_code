"""Excel/workbook tools. Deterministic: pinned locale, explicit paths."""

from pathlib import Path
from typing import Any

# Optional: use openpyxl/pandas when available
try:
    import openpyxl
except ImportError:
    openpyxl = None

_WORKBOOK_STACK: list[Any] = []  # Each entry: (wb, path) when openpyxl; else {"path": path, "open": True}


def _normalize_range(range_spec: str) -> str:
    """Turn A1G999 into A1:G999 for openpyxl."""
    s = range_spec.strip().upper()
    if ":" in s:
        return s
    if len(s) >= 2 and s[0].isalpha() and s[1].isdigit():
        i = 1
        while i < len(s) and s[i].isdigit():
            i += 1
        if i < len(s):
            return s[:i] + ":" + s[i:]
    return s


def excel_open_workbook(path: str, **kwargs: Any) -> Any:
    """Open a workbook and push onto stack; read_table uses the top (most recent)."""
    global _WORKBOOK_STACK
    if openpyxl is None:
        entry = {"path": path, "open": True}
        _WORKBOOK_STACK.append(entry)
        return entry
    p = Path(path)
    if not p.is_absolute():
        p = Path.cwd() / p
    wb = openpyxl.load_workbook(p, read_only=False, data_only=True)
    entry = (wb, str(p))
    _WORKBOOK_STACK.append(entry)
    return entry


def excel_read_table(
    sheet: str,
    range: str | None = None,
    range_spec: str | None = None,
    **kwargs: Any,
) -> Any:
    """Read a range as a table (list of dicts). Uses top of workbook stack (most recently opened)."""
    global _WORKBOOK_STACK
    r = range_spec or range or ""
    if not _WORKBOOK_STACK:
        if openpyxl:
            raise RuntimeError("No workbook open. Use Open workbook first.")
        return []
    wb = _WORKBOOK_STACK[-1]
    if openpyxl is None:
        return []
    wb_obj, _ = wb if isinstance(wb, tuple) else (None, None)
    if wb_obj is None:
        return []
    ws = wb_obj[sheet]
    r = _normalize_range(r)
    rows = list(ws[r])
    if not rows:
        return []
    headers = [str(c.value) if c.value is not None else f"_c{i}" for i, c in enumerate(rows[0])]
    out = []
    for row in rows[1:]:
        out.append(dict(zip(headers, (c.value for c in row))))
    return out


def excel_export(source: Any, path: str, **kwargs: Any) -> None:
    """Export table (list of dicts) to CSV or XLSX."""
    if not source or not isinstance(source, list):
        return
    p = Path(path)
    if not p.is_absolute():
        p = Path.cwd() / p
    p.parent.mkdir(parents=True, exist_ok=True)
    if p.suffix.lower() == ".csv":
        import csv
        with open(p, "w", newline="", encoding="utf-8") as f:
            if source and isinstance(source[0], dict):
                w = csv.DictWriter(f, fieldnames=source[0].keys())
                w.writeheader()
                w.writerows(source)
        return
    if openpyxl is None:
        raise RuntimeError("XLSX export requires openpyxl. Install with: pip install openpyxl")
    wb = openpyxl.Workbook()
    ws = wb.active
    if source and isinstance(source[0], dict):
        headers = list(source[0].keys())
        ws.append(headers)
        for row in source:
            ws.append([row.get(h) for h in headers])
    wb.save(p)


def table_add_column(table: Any, name: str, expr: Any, **kwargs: Any) -> Any:
    """Add a column to the table. expr can be a constant (same for all rows)."""
    if not isinstance(table, list) or not table:
        return table
    if not isinstance(table[0], dict):
        return table
    for row in table:
        row[name] = expr
    return table


def _numeric_value(v: Any) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, dict) and "amount" in v:
        return float(v["amount"])
    if isinstance(v, dict) and "value" in v:
        return float(v["value"])
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _is_numeric(v: Any) -> bool:
    """Return True if v can meaningfully be treated as a number."""
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return True
    if isinstance(v, dict) and ("amount" in v or "value" in v):
        return True
    if isinstance(v, str):
        try:
            float(v)
            return True
        except ValueError:
            return False
    return False


def _truthy(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v != 0
    if isinstance(v, str):
        return len(v.strip()) > 0
    return True


def table_filter(table: Any, condition: Any, **kwargs: Any) -> Any:
    """Filter rows by a condition. condition: {type: 'comparison', ...} or {type: 'not', expr: ...}."""
    if not isinstance(table, list) or not table:
        return table
    if not condition or not isinstance(condition, dict):
        return table
    if condition.get("type") == "not":
        inner = condition.get("expr")
        if not isinstance(inner, dict):
            return table
        field = None
        if inner.get("type") == "qualified" and inner.get("field"):
            field = inner["field"]
        elif inner.get("type") == "ref" and inner.get("name"):
            field = inner["name"]
        if field is None:
            return table
        return [row for row in table if isinstance(row, dict) and not _truthy(row.get(field))]
    if condition.get("type") != "comparison":
        return table
    left = condition.get("left")
    op = condition.get("op")
    right = condition.get("right")
    if op is None or left is None:
        return table
    if isinstance(left, dict) and left.get("type") == "qualified":
        field = left.get("field")
    else:
        return table
    # Resolve the right-hand value from IR expression dicts
    if isinstance(right, dict):
        if right.get("type") == "string":
            right_raw = right.get("value")
        elif right.get("type") == "number":
            right_raw = right.get("value")
        elif right.get("type") == "money":
            right_raw = right.get("amount")
        else:
            right_raw = right.get("value", right)
    else:
        right_raw = right
    out = []
    for row in table:
        if not isinstance(row, dict):
            continue
        val = row.get(field)
        if op == ">":
            if _numeric_value(val) > _numeric_value(right_raw):
                out.append(row)
        elif op == ">=":
            if _numeric_value(val) >= _numeric_value(right_raw):
                out.append(row)
        elif op == "<":
            if _numeric_value(val) < _numeric_value(right_raw):
                out.append(row)
        elif op == "<=":
            if _numeric_value(val) <= _numeric_value(right_raw):
                out.append(row)
        elif op == "=":
            if val == right_raw:
                out.append(row)
            elif _is_numeric(val) and _is_numeric(right_raw) and _numeric_value(val) == _numeric_value(right_raw):
                out.append(row)
        elif op == "!=":
            if val == right_raw:
                continue
            if _is_numeric(val) and _is_numeric(right_raw) and _numeric_value(val) == _numeric_value(right_raw):
                continue
            out.append(row)
        else:
            out.append(row)
    return out


def _sort_key(val: Any) -> tuple[int, str]:
    """Return a comparable key so None and mixed types sort without errors. Nones first, then by string repr."""
    if val is None:
        return (0, "")
    return (1, str(val))


def table_sort(table: Any, by: Any, ascending: bool = True, **kwargs: Any) -> Any:
    """Sort table (list of dicts) by a column. by: field name (string) or resolved qualified ref dict with 'field'."""
    if not isinstance(table, list) or not table:
        return table
    if not isinstance(table[0], dict):
        return table
    key = None
    if isinstance(by, str):
        key = by
    elif isinstance(by, dict) and by.get("type") == "qualified" and by.get("field"):
        key = by["field"]
    if key is None:
        return table
    return sorted(
        table,
        key=lambda row: _sort_key(row.get(key)),
        reverse=not ascending,
    )
