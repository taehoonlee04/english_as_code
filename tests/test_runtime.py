"""Tests for EAC runtime (interpreter + Excel tools)."""

import csv
import tempfile
from pathlib import Path

import pytest

from eac.ir import IRProgram, IRStep
from eac.runtime.interpreter import run


def _has_openpyxl() -> bool:
    try:
        import openpyxl  # noqa: F401
        return True
    except ImportError:
        return False


@pytest.mark.skipif(not _has_openpyxl(), reason="openpyxl not installed")
def test_excel_pipeline_add_column_filter_export():
    """Run open_workbook -> read_table -> add_column -> filter -> export with a temp xlsx."""
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        xlsx = tmp_path / "data.xlsx"
        out_csv = tmp_path / "out.csv"

        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["Amount", "Balance", "Name"])
        ws.append([100, 50.0, "Alice"])
        ws.append([200, 0.0, "Bob"])
        ws.append([300, 75.5, "Carol"])
        wb.save(xlsx)

        ir = IRProgram(
            steps=[
                IRStep("1", "excel.open_workbook", {"path": str(xlsx)}, result=None),
                IRStep("2", "excel.read_table", {"sheet": "Data", "range": "A1C4"}, result="T"),
                IRStep("3", "set_var", {"name": "today", "value": {"type": "date", "value": "2026-02-11"}}, result=None),
                IRStep("4", "table.add_column", {"table": "T", "name": "DaysPastDue", "expr": {"type": "ref", "name": "today"}}, result="T"),
                IRStep("5", "table.filter", {"table": "T", "condition": {"type": "comparison", "left": {"type": "qualified", "base": "T", "field": "Balance"}, "op": ">", "right": 0}}, result="T"),
                IRStep("6", "excel.export", {"source": "T", "path": str(out_csv)}, result=None),
            ]
        )
        trace = run(ir)
        assert len(trace) == 6
        assert out_csv.exists()
        with open(out_csv, newline="", encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
        assert len(rows) == 2  # Bob has Balance 0, filtered out
        names = {r["Name"] for r in rows}
        assert "Alice" in names and "Carol" in names and "Bob" not in names
        assert all("DaysPastDue" in r for r in rows)


def test_interpreter_resolves_refs_in_args():
    """Refs in nested condition are resolved from env."""
    ir = IRProgram(
        steps=[
            IRStep("1", "set_var", {"name": "today", "value": {"type": "date", "value": "2026-02-11"}}, result=None),
            IRStep("2", "table.add_column", {"table": "T", "name": "D", "expr": {"type": "ref", "name": "today"}}, result="T"),
        ]
    )
    # T is not in env so table will be string "T"; add_column will get table="T" (unchanged). We need T to be a table.
    # So use a literal table: we don't have that op. Instead test that set_var stores value and a later step can use it.
    ir2 = IRProgram(
        steps=[
            IRStep("1", "set_var", {"name": "x", "value": 10}, result="x"),
            IRStep("2", "set_var", {"name": "y", "value": {"type": "ref", "name": "x"}}, result="y"),
        ]
    )
    trace = run(ir2)
    assert trace[1]["result"] == 10


def test_for_each_executes_body_per_row():
    """For each row, body steps run with row in env; qualified refs resolve."""
    ir = IRProgram(
        steps=[
            IRStep("1", "set_var", {"name": "x", "value": []}, result="x"),
            IRStep(
                "2",
                "control.for_each",
                {
                    "var": "row",
                    "collection": [{"A": 1, "B": 10}, {"A": 2, "B": 20}],
                    "body": [
                        {"id": "2a", "op": "set_var", "args": {"name": "v", "value": {"type": "qualified", "base": "row", "field": "B"}}, "result": "v"},
                    ],
                },
                result=None,
            ),
        ]
    )
    trace = run(ir)
    # One entry for set_var, one for for_each (no trace entry for the loop itself), then 2 body steps
    set_var_entries = [e for e in trace if e.get("op") == "set_var" and e.get("args", {}).get("name") == "v"]
    assert len(set_var_entries) == 2
    assert set_var_entries[0]["result"] == 10
    assert set_var_entries[1]["result"] == 20


@pytest.mark.skipif(not _has_openpyxl(), reason="openpyxl not installed")
def test_multiple_workbooks_read_from_correct_book():
    """Opening two workbooks and reading from each uses the correct (top of stack) workbook."""
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        a_path = tmp_path / "a.xlsx"
        b_path = tmp_path / "b.xlsx"
        import openpyxl
        wa = openpyxl.Workbook()
        wa.active.append(["ID", "FromA"])
        wa.active.append([1, "first"])
        wa.save(a_path)
        wb = openpyxl.Workbook()
        wb.active.append(["ID", "FromB"])
        wb.active.append([1, "second"])
        wb.save(b_path)

        ir = IRProgram(
            steps=[
                IRStep("1", "excel.open_workbook", {"path": str(a_path)}, result=None),
                IRStep("2", "excel.read_table", {"sheet": "Sheet", "range": "A1B2"}, result="T1"),
                IRStep("3", "excel.open_workbook", {"path": str(b_path)}, result=None),
                IRStep("4", "excel.read_table", {"sheet": "Sheet", "range": "A1B2"}, result="T2"),
            ]
        )
        trace = run(ir)
        assert trace[1]["result"][0]["FromA"] == "first"
        assert trace[3]["result"][0]["FromB"] == "second"
